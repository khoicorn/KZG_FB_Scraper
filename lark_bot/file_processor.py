import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image
import requests
import logging
from typing import Optional, Tuple
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

from zipfile import ZipFile, ZIP_DEFLATED
import hashlib
from urllib.parse import urlparse
import os

class ExcelImageExporter:
    """Optimized Excel exporter with parallel image processing."""
    """Optimized Excel exporter with parallel image processing."""

    def __init__(self, 
                    image_size: Tuple[int, int] = (80, 80),
                    row_height: int = 80,
                    image_col_width: int = 18,
                    timeout: int = 10,
                    max_workers: int = 10):
        """
        Initialize the Excel exporter.
        
        Args:
            image_size: Tuple of (width, height) for image thumbnails
            row_height: Excel row height for image rows
            image_col_width: Width of the image column
            timeout: Request timeout in seconds
            max_workers: Max threads for parallel image downloads
        """
        self.image_size = image_size
        self.row_height = row_height
        self.image_col_width = image_col_width
        self.timeout = timeout
        self.max_workers = max_workers
        self.logger = logging.getLogger(__name__)

    def _download_and_process_image(self, url: str) -> Optional[bytes]:
        """
        Download and process an image from URL.
        """
        try:
            headers = {'User-Agent': 'Mozilla/5.0'}
            response = requests.get(url, headers=headers, timeout=self.timeout)
            response.raise_for_status()
            
            with Image.open(BytesIO(response.content)) as img:
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                img.thumbnail(self.image_size, Image.Resampling.LANCZOS)
                
                with BytesIO() as buffer:
                    img.save(buffer, format="PNG", optimize=True)
                    return buffer.getvalue()
                    
        except Exception as e:
            self.logger.warning(f"Image processing failed for {url}: {str(e)}")
            return None

    def _setup_header_styling(self, ws, num_cols: int):
        """Apply styling to header row."""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # --- MODIFIED --- 
        # The number of columns is now based on the final list of headers, not the original df
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

    def _auto_adjust_column_widths(self, ws, df: pd.DataFrame):
        """Auto-adjust column widths based on content."""
        # --- MODIFIED ---
        # This function now uses the final column order for adjusting width.
        # We also need to skip the 'Image' column which has a fixed width.
        for col_idx, column in enumerate(df.columns, 1):
            if column == 'Image':
                continue
            
            # Check if the column exists in the DataFrame before trying to access it
            if column in df:
                max_length = max(
                    len(str(column)),
                    df[column].astype(str).str.len().max()
                )
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    def export_to_excel(self, 
                        df: pd.DataFrame, 
                        image_column: str) -> BytesIO:
        """
        Export DataFrame to Excel with images (in-memory only).
        """
        if image_column not in df.columns:
            raise ValueError(f"Image column '{image_column}' not found in DataFrame")
        
        df = df.reset_index(drop=True).copy()
        if "No" not in df.columns:
            df.insert(0, "No", range(1, len(df) + 1))

        # --- SIMPLIFIED COLUMN SETUP ---
        # Start with the DataFrame's original columns
        final_excel_columns = list(df.columns)
        # Add the 'Image' column at the end
        final_excel_columns.append('Image')

        wb = Workbook()
        ws = wb.active
        ws.title = "Data with Images"
        
        # --- MODIFIED ---
        # Write headers based on our new `final_excel_columns` list.
        for col_idx, col_name in enumerate(final_excel_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # --- REMOVE ---
        # We no longer need this part because the 'Image' header is already in our list.
        # image_col_idx = len(df.columns) + 1
        # ws.cell(row=1, column=image_col_idx, value="Image")
        
        # Apply header styling based on the final number of columns
        self._setup_header_styling(ws, len(final_excel_columns))
        
        # Phase 1: Write data and collect image URLs
        download_tasks = {}
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            # --- MODIFIED ---
            # Write data cell by cell from the original 'row' object,
            # but into the correct column positions defined by 'final_excel_columns'.
            for col_idx, col_name in enumerate(final_excel_columns, 1):
                if col_name != 'Image' and col_name in row:
                    ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            
            if pd.notna(row[image_column]) and str(row[image_column]).strip():
                download_tasks[row_idx] = str(row[image_column])

        # Phase 2: Parallel image downloads (This part is unchanged)
        image_data = {}
        if download_tasks:
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_row = {
                    executor.submit(self._download_and_process_image, url): row_idx
                    for row_idx, url in download_tasks.items()
                }
                
                for future in as_completed(future_to_row):
                    row_idx = future_to_row[future]
                    image_data[row_idx] = future.result()

        # --- ADD ---
        # Get the new 1-based index for the 'Image' column for placing images.
        image_col_idx = final_excel_columns.index('Image') + 1

        # Phase 3: Insert images into worksheet
        successful_images = 0
        failed_images = 0
        for row_idx, img_bytes in image_data.items():
            if img_bytes:
                try:
                    img_buffer = BytesIO(img_bytes)
                    img = OpenPyxlImage(img_buffer)
                    
                    # --- MODIFIED ---
                    # The anchor position now correctly uses the new image_col_idx.
                    image_col_letter = get_column_letter(image_col_idx)
                    img.anchor = f"{image_col_letter}{row_idx}"
                    
                    ws.add_image(img)
                    ws.row_dimensions[row_idx].height = self.row_height
                    successful_images += 1
                    
                except Exception as e:
                    self.logger.error(f"Failed to add image to Excel for row {row_idx}: {e}")
                    failed_images += 1
            else:
                failed_images += 1
        

        # --- SIMPLIFIED ---
        # Auto-adjust column widths for the entire DataFrame.
        self._auto_adjust_column_widths(ws, df)
        
        # Process hyperlinks (This logic remains largely the same but uses the final column indices)
        hyperlink_columns = ["destination_url", "ad_url", "thumbnail_url"]
        for col_name in hyperlink_columns:
            if col_name in final_excel_columns:
                # --- MODIFIED ---
                # Get index from the final list
                col_idx = final_excel_columns.index(col_name) + 1
                for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                    url = row.get(col_name) # Use .get() for safety
                    if pd.notna(url) and str(url).strip():
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = "Click here"
                        cell.hyperlink = str(url)
                        cell.font = Font(color="0563C1", underline="single")
                
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

        # Set image column width
        image_col_letter = get_column_letter(image_col_idx)
        ws.column_dimensions[image_col_letter].width = self.image_col_width

        self.logger.info(f"Export completed: {successful_images} images added, {failed_images} failed")
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

def _filename_from_url(url: str, prefix: str) -> str:
    """
    Create a stable filename from URL, preserving extension when possible.
    """
    parsed = urlparse(url)
    # Try to keep the last path segment’s extension
    base = os.path.basename(parsed.path) or "file"
    name, ext = os.path.splitext(base)
    if not ext:
        ext = ".bin"
    # Stable short hash to avoid collisions
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:10]
    safe_prefix = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in prefix)[:30]
    return f"{safe_prefix}{ext}"

def _download_bytes(url: str, timeout: int = 20) -> bytes | None:
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=timeout, stream=True)
        r.raise_for_status()
        # Avoid massive memory for huge streams; cap to ~200MB each file just in case
        return r.content[:200 * 1024 * 1024]
    except Exception:
        return None

def build_media_zip(
     df: pd.DataFrame,
    col: str,
    zip_basename_prefix: str,
    max_workers: int = 8,
    max_zip_bytes: int = 28 * 1024 * 1024,  # ~28MB safe under 30MB
) -> list[tuple[str, BytesIO]]:
    if col not in df.columns:
        return []

    # Collect (No, url) pairs
    rows = []
    for _, row in df.iterrows():
        val = str(row[col]).strip() if pd.notna(row[col]) else ""
        if val:
            no_val = None
            if "No" in df.columns:
                try:
                    no_val = int(row["No"])
                except Exception:
                    no_val = None
            rows.append((no_val, val))

    # Deduplicate by URL string
    seen = set()
    unique_rows = []
    for no_val, u in rows:
        if u not in seen:
            seen.add(u)
            unique_rows.append((no_val, u))

    # Parallel download
    blobs: list[tuple[int | None, str, bytes | None]] = []
    if unique_rows:
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            futs = {ex.submit(_download_bytes, u): (no_val, u) for no_val, u in unique_rows}
            for fut in as_completed(futs):
                no_val, u = futs[fut]
                data = fut.result()
                blobs.append((no_val, u, data))

    parts: list[tuple[str, BytesIO]] = []
    if not blobs:
        return parts

    part_idx = 1
    current = BytesIO()
    zf = ZipFile(current, "w", compression=ZIP_DEFLATED)
    written_bytes = 0
    manifest_rows = []

    def _close_and_push():
        nonlocal current, zf, manifest_rows, written_bytes, part_idx, parts
        zf.writestr("manifest.csv", "No,column,url\n".encode("utf-8") +
                    "\n".join([f"{no},{col},{u}" for no, u in manifest_rows]).encode("utf-8"))
        zf.close()
        current.seek(0)
        fname = f"{zip_basename_prefix}_{col}_media_part{part_idx}.zip"
        parts.append((fname, current))
        part_idx += 1
        current = BytesIO()
        zf = ZipFile(current, "w", compression=ZIP_DEFLATED)
        manifest_rows = []
        written_bytes = 0

    for no_val, u, data in blobs:
        if not data:
            continue
        # build filename
        base_fname = _filename_from_url(u, prefix=col)
        if no_val is not None:
            fname = f"{no_val}_{base_fname}"
        else:
            fname = base_fname

        estimated_added = len(data) + 2048
        if written_bytes + estimated_added > max_zip_bytes and written_bytes > 0:
            _close_and_push()
        zf.writestr(fname, data)
        manifest_rows.append((no_val, u))
        written_bytes += estimated_added

    if written_bytes > 0 or not parts:
        _close_and_push()

    return parts


def export_dataframe_with_images(df: pd.DataFrame, 
                                image_column: str,
                                **kwargs) -> BytesIO:
    """
    Convenience function to export DataFrame with images to Excel.
    
    Args:
        df: DataFrame to export
        image_column: Column name containing image URLs
        **kwargs: Additional arguments for ExcelImageExporter
        
    Returns:
        BytesIO object containing the Excel file
    """
    exporter = ExcelImageExporter(**kwargs)
    return exporter.export_to_excel(df, image_column)

def generate_excel_report(crawler):
    """Generate Excel report from crawler data with robust error handling."""
    today = datetime.now().strftime("%Y-%m-%d")
    time.sleep(1)  # Reduced sleep time
    
    # Process crawler data
    crawler.start()
    
    # Wait for processing to complete
    while crawler.queue_manager.get_queue_position(crawler.chat_id) is not None:
        time.sleep(0.5)
    
    # crawler.data_to_dataframe()

    if crawler.df.empty:
        return None, f"{crawler.keyword.replace('.', '-')}_{today}_results.xlsx", crawler.df

    # Create exporter with optimized settings
    try:
        exporter = ExcelImageExporter(
            image_size=(100, 100),
            row_height=100,
            timeout=15,
            max_workers=10
        )
        
        excel_buffer = exporter.export_to_excel(
            df=crawler.df,
            image_column='thumbnail_url'
        )
        return excel_buffer, f"{crawler.keyword.replace('.', '-')}_{today}_results.xlsx", crawler.df
    except Exception as e:
        logging.error(f"Excel generation failed: {str(e)}")
        return None, f"{crawler.keyword.replace('.', '-')}_{today}_results.xlsx", crawler.df
    finally:
        # Ensure crawler resources are cleaned up
        if hasattr(crawler, 'driver') and crawler.driver:
            try:
                crawler.driver.quit()
            except:
                pass